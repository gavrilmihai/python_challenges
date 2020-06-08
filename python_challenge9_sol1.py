'''Using import openpyxl and from openpyxl import Workbook to move data into & out of xlsx sheets.
Take the data in the device dictionary & add this to an excel file with the keys as headers & items as the sub-fields.
Modify some of the xls data (use python if you can, else for time just do it manually)\n
Import the data back into python & display it to the terminal.
'''
import openpyxl
import json
device_dictionary = {'device':['csr1kv', 'asav','vios'],'hostname':['r1-csr1kv','f1-asav','r2-vios'],'domain_name':['py.init','py.init','py.init'],'username':['squirrel1','squirrel2','squirrel3'],'password':['secretsquirrel43','secretsquirrel42','secretsquirrel9'],'ip_address':['192.168.1.1','192.168.22.22','192.168.33.33'],'ip_netmask':['255.255.255.0','255.255.0.0','255.0.0.0']}
wb = openpyxl.load_workbook('day9.xlsx')
sheet=wb.active
#sheet.title='Day9-Sheet1'
#wb.remove(wb.get_sheet_by_name('Sheet2')) 
wb.create_sheet('Day9-Insert-Dict') #create a new Sheet for inserting Dictionary data 
sheet=wb['Day9-Insert-Dict']
key_columns='F'  #Start in Column B row 4
row_index=10
for k,list in device_dictionary.items():
    sheet[key_columns+str(row_index)]= k   #Update Cell Value with Key
    value_columns=chr(ord(key_columns)+1)  #Move Column to right stay in same row 
    for value in list:                     #As the value of Device_dictionary is a list , iterating over it 
        sheet[value_columns+str(row_index)]= value #Update Cell Value with Key 
        value_columns=chr(ord(value_columns)+1)    #Move Column to right stay in same row 
    row_index+=1                          #Move to row bellow 
wb.save('day9.xlsx')
wb = openpyxl.load_workbook('day9.xlsx') #reopen file and read data into dictionary
sheet=wb['Day9-Sheet1']
key_columns='A'
key_row=1  #Specify the key element for disctionary 
keys=[]
values=[]
while sheet[key_columns+str(key_row)].value != None :
    keys.append(sheet[key_columns+str(key_row)].value)
    list=[]  
    while sheet[key_columns+str(key_row+1)].value != None :
        list.append(sheet[key_columns+str(key_row+1)].value)
        key_row+=1
    values.append(list)
    key_row=1
    key_columns=chr(ord(key_columns)+1)  #Move Column to right stay in same row 
dictionary = dict(zip(keys, values))
print(dictionary)
my_json=json.dumps(dictionary)
print(my_json)