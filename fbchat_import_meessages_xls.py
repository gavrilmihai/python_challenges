from fbchat import Client
from fbchat.models import *
import openpyxl
import json
wb = openpyxl.Workbook() 
wb.create_sheet('FB-Messages')
sheet=wb['FB-Messages']
client = Client("gavrilmihai@gmail.com", "^&Tytan76")
print("Own id: {}".format(client.uid))
# Fetches a list of the 20 top threads you're currently chatting with
threads = client.fetchThreadList() 
# Fetches the next 20 threads
threads += client.fetchThreadList(limit=20)
threads += client.fetchThreadList(limit=20)
threads += client.fetchThreadList(limit=20)
threads += client.fetchThreadList(limit=20)
#this is last 100 threads for each will retrieve last 20 messages
for t in range(0,len(threads)): 
    thread_name=threads[t].__dict__['name']
    thread_id=threads[t].__dict__['uid']
    wb.create_sheet(thread_id)
    sheet=wb[thread_id]
    messages = client.fetchThreadMessages(thread_id=thread_id, limit=20)
    sheet['A1']= "Index"
    sheet['B1']= "Text"
    sheet['C1']= "Mentions"
    sheet['D1']= "Emoji_size"
    sheet['E1']= "Uid.mid"
    sheet['D1']= "Author"
    sheet['E1']= "Timestamp"
    sheet['F1']= "Is Read"
    sheet['F1']= "Read By"
    value_rows=2
    for i in range(0,len(messages)):
        value_columns='A'
        row= value_rows + i
        sheet[value_columns+str(row)]= i+1
        message=messages[i].__dict__
        for k,v in message.items():
            value_columns=chr(ord(value_columns)+1)  #Move Column to right stay in same row 
            print(value_columns+str(row))
            sheet[value_columns+str(row)]= str(v)

wb.save('facebook_messages.xlsx')